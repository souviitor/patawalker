from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_required, current_user
from functools import wraps
from models.models import db, User, ServicePlan, WalkDuration, ServiceOrder, FinancialRecord, AppSetting, PlanType, UserRole
from datetime import datetime, date
import io

admin_bp = Blueprint('admin', __name__)

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Acesso restrito ao administrador.', 'error')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return login_required(decorated)

@admin_bp.route('/dashboard')
@admin_required
def dashboard():
    total_users = User.query.filter_by(role=UserRole.TUTOR).count()
    total_orders = ServiceOrder.query.count()
    
    from sqlalchemy import func
    revenue = db.session.query(func.sum(ServiceOrder.amount))\
        .filter(ServiceOrder.payment_status == 'APPROVED').scalar() or 0
    
    pending = ServiceOrder.query.filter_by(status='PENDING').count()
    
    recent_users = User.query.order_by(User.created_at.desc()).limit(10).all()
    recent_orders = ServiceOrder.query.order_by(ServiceOrder.created_at.desc()).limit(10).all()
    
    return render_template('admin/dashboard.html',
        total_users=total_users, total_orders=total_orders,
        revenue=revenue, pending=pending,
        recent_users=recent_users, recent_orders=recent_orders)

@admin_bp.route('/reports')
@admin_required
def reports():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = User.query.filter(User.role != UserRole.ADMIN)
    order_query = ServiceOrder.query
    
    if start_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        query = query.filter(User.created_at >= start)
        order_query = order_query.filter(ServiceOrder.created_at >= start)
    if end_date:
        end = datetime.strptime(end_date, '%Y-%m-%d')
        query = query.filter(User.created_at <= end)
        order_query = order_query.filter(ServiceOrder.created_at <= end)
    
    users = query.all()
    orders = order_query.all()
    
    from sqlalchemy import func
    revenue = db.session.query(func.sum(ServiceOrder.amount))\
        .filter(ServiceOrder.payment_status == 'APPROVED').scalar() or 0
    
    return render_template('admin/reports.html',
        users=users, orders=orders, revenue=revenue,
        start_date=start_date, end_date=end_date)

@admin_bp.route('/export-excel')
@admin_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Instale openpyxl para exportar Excel.', 'error')
        return redirect(url_for('admin.reports'))
    
    wb = openpyxl.Workbook()
    
    # Users sheet
    ws1 = wb.active
    ws1.title = "Cadastros"
    ws1.append(["ID", "Nome", "Email", "Documento", "Cidade", "Data Cadastro"])
    users = User.query.filter(User.role != UserRole.ADMIN).all()
    for u in users:
        ws1.append([u.id, u.full_name, u.email, u.document, u.city, str(u.created_at)])
    
    # Orders sheet
    ws2 = wb.create_sheet("Pedidos")
    ws2.append(["ID", "Tutor", "Pet", "Serviço", "Status", "Valor", "Data"])
    orders = ServiceOrder.query.all()
    for o in orders:
        ws2.append([o.id, o.tutor.full_name if o.tutor else '', 
                    o.pet.name if o.pet else '', o.service_type,
                    o.status.value if o.status else '', 
                    float(o.amount) if o.amount else 0,
                    str(o.created_at)])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'relatorio_patawalker_{date.today()}.xlsx')

@admin_bp.route('/service-rates', methods=['GET', 'POST'])
@admin_required
def service_rates():
    plans = ServicePlan.query.all()
    durations = WalkDuration.query.all()
    return render_template('admin/service_rates.html', plans=plans, durations=durations)

@admin_bp.route('/plan/save', methods=['POST'])
@admin_required
def save_plan():
    plan_id = request.form.get('plan_id')
    plan_type = request.form.get('plan_type')
    name = request.form.get('name')
    description = request.form.get('description')
    walks = int(request.form.get('walks_per_week', 2))
    duration = int(request.form.get('duration_minutes', 30))
    price = float(request.form.get('price', 0))
    
    if plan_id:
        plan = ServicePlan.query.get(plan_id)
    else:
        plan = ServicePlan()
        db.session.add(plan)
    
    plan.plan_type = PlanType[plan_type]
    plan.name = name
    plan.description = description
    plan.walks_per_week = walks
    plan.duration_minutes = duration
    plan.price = price
    db.session.commit()
    flash('Plano salvo com sucesso!', 'success')
    return redirect(url_for('admin.service_rates'))

@admin_bp.route('/duration/save', methods=['POST'])
@admin_required
def save_duration():
    dur_id = request.form.get('duration_id')
    name = request.form.get('name')
    minutes = int(request.form.get('duration_minutes', 30))
    price = float(request.form.get('price', 0))
    
    if dur_id:
        dur = WalkDuration.query.get(dur_id)
    else:
        dur = WalkDuration()
        db.session.add(dur)
    
    dur.name = name
    dur.duration_minutes = minutes
    dur.price = price
    db.session.commit()
    flash('Duração salva com sucesso!', 'success')
    return redirect(url_for('admin.service_rates'))

@admin_bp.route('/financial')
@admin_required
def financial():
    from sqlalchemy import func
    records = FinancialRecord.query.order_by(FinancialRecord.date.desc()).limit(100).all()
    total_income = db.session.query(func.sum(FinancialRecord.amount))\
        .filter(FinancialRecord.type == 'income').scalar() or 0
    total_expense = db.session.query(func.sum(FinancialRecord.amount))\
        .filter(FinancialRecord.type == 'expense').scalar() or 0
    total_refund = db.session.query(func.sum(FinancialRecord.amount))\
        .filter(FinancialRecord.type == 'refund').scalar() or 0
    
    return render_template('admin/financial.html', 
        records=records, total_income=total_income,
        total_expense=total_expense, total_refund=total_refund)

@admin_bp.route('/users')
@admin_required  
def users():
    all_users = User.query.filter(User.role != UserRole.ADMIN)\
        .order_by(User.created_at.desc()).all()
    return render_template('admin/users.html', users=all_users)